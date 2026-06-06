#!/usr/bin/env python3
"""Genera ConfigPackage_Payroll.xlsx para Business Central."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(__file__), 'ConfigPackage_Payroll.xlsx')
VIG = '01/01/2025'

TIPO_MAP = {
    'HR': 'Haber Remunerativo', 'HN': 'Haber No Remunerativo',
    'DE': 'Descuento Empleado', 'CP': 'Contribucion Patronal',
    'RE': 'Retencion', 'SS': 'Seguridad Social',
}
PRIMARY_ACC = {
    'HR': 'REMUNERATIVO_BRUTO', 'HN': 'NO_REMUNERATIVO',
    'DE': 'TOTAL_DESCUENTOS',   'RE': 'TOTAL_DESCUENTOS',
    'SS': 'TOTAL_DESCUENTOS',   'CP': 'TOTAL_CONTRIBUCIONES',
}
ACC_NAMES = ['BASE_SS','BASE_OS','BASE_SINDICAL','BASE_LRT','BASE_IG4',
             'BASE_SAC','BASE_PROMEDIO','BASE_FERIADO','BASE_ZONA','BASE_AUSENTISMO']
CCT_CODES = ['175/75','768/19','729/15','ESP','130/75','372/04','ADM']

# ── Acumuladores ────────────────────────────────────────────────────
ACUMULADORES = [
    ('REMUNERATIVO_BRUTO','Acum: Total Remunerativo','Haber Remunerativo'),
    ('NO_REMUNERATIVO','Acum: Total No Remunerativo','Haber No Remunerativo'),
    ('TOTAL_DESCUENTOS','Acum: Total Descuentos','Descuento Empleado'),
    ('TOTAL_CONTRIBUCIONES','Acum: Total Contribuciones','Contribucion Patronal'),
    ('BASE_SS','Acum: Base Seguridad Social','Haber Remunerativo'),
    ('BASE_OS','Acum: Base Obra Social','Haber Remunerativo'),
    ('BASE_SINDICAL','Acum: Base Sindical','Haber Remunerativo'),
    ('BASE_LRT','Acum: Base LRT','Haber Remunerativo'),
    ('BASE_IG4','Acum: Base Imp. Ganancias 4ta Cat.','Haber Remunerativo'),
    ('BASE_SAC','Acum: Base SAC / Aguinaldo','Haber Remunerativo'),
    ('BASE_PROMEDIO','Acum: Base Promedio','Haber Remunerativo'),
    ('BASE_FERIADO','Acum: Base Feriado Embarcado','Haber Remunerativo'),
    ('BASE_ZONA','Acum: Base Zona Desfavorable','Haber Remunerativo'),
    ('BASE_AUSENTISMO','Acum: Base Ausentismo','Haber Remunerativo'),
]

# ── Conceptos ────────────────────────────────────────────────────────
# (cod, nombre, tipo, aplica_tipo_liq, ord, activo,
#  cct[7]: 175,768,729,ESP,130,372,ADM,
#  acc[10]: SS,OS,SIND,LRT,IG4,SAC,PROM,FER,ZONA,AUSEN, formula_hint)
CONCEPTOS = [
    # SUELDOS EMBARCADOS
    ('1003','Sueldo','HR',' ',10,'TRUE', 1,1,1,0,1,1,1, 1,1,1,1,1,1,1,0,1,1,'BASICO * PCT_ESCALA'),
    ('1013','Sueldo de navegacion','HR',' ',15,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1023','Sueldo de feriado','HR',' ',20,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1033','Sueldo de dique','HR',' ',25,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1043','Sueldo de pilotaje','HR',' ',30,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1045','Servicio de asistencia y/o remolque','HR',' ',35,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1053','Antiguedad','HR',' ',50,'TRUE', 1,1,1,0,1,1,0, 1,1,1,1,1,1,1,1,0,1,'BASICO * PCT_ESCALA * ANIOS_ANTIGUEDAD * PCT_ANTIG'),
    ('1063','Sueldo de franco','HR',' ',60,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1073','Sueldo a ordenes','HR',' ',65,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1075','Articulo 19 CCT 729/2015','HR',' ',70,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1076','Articulo 48 Parrafo 11 CCT 729/2015','HR',' ',75,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1083','Sueldo de puerto','HR',' ',80,'FALSE', 1,1,1,1,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('1093','Gratificacion produccion langostino colas','HR',' ',90,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('1103','Gratificacion trabajos a bordo','HR',' ',95,'FALSE', 1,1,1,1,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1113','Sueldo basico articulo 16 AACPyPP','HR',' ',100,'TRUE', 0,1,0,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1153','Ajuste de dias de puerto','HR',' ',105,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    # NO REMUNERATIVOS
    ('1173','Adicional bodega no remunerativo','HN',' ',200,'TRUE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('1174','Adicional bodega remunerativo','HR',' ',205,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1183','Ajuste de antiguedad','HR',' ',210,'FALSE', 1,1,1,0,1,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    # PRODUCCION OFICIALES
    ('1203','Produccion bruta merluza','HR',' ',110,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1213','Produccion bruta calamar entero','HR',' ',115,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1223','Produccion bruta langostino entero','HR',' ',120,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1233','Produccion bruta langostino colas','HR',' ',125,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1243','Produccion bruta calamar vaina','HR',' ',130,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1253','Produccion bruta calamar rejos','HR',' ',135,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1263','Produccion bruta langostino cola rota','HR',' ',140,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    # PRODUCCION MARINERIA
    ('1333','Produccion merluza','HR',' ',110,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('1343','Produccion calamar entero','HR',' ',115,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1373','Produccion calamar vaina','HR',' ',130,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1383','Produccion calamar rejos','HR',' ',135,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1387','Complemento no rem produccion calamar','HN',' ',220,'TRUE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('1388','Complemento no rem produccion langostino','HN',' ',225,'TRUE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('1393','Produccion langostino cola rota','HR',' ',140,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('1413','A cuenta futuros aumentos (729/15)','HN',' ',230,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1513','Ajuste a cuenta futuros aumentos','HR',' ',235,'FALSE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1533','Adicional articulo 23','HR',' ',145,'TRUE', 1,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1534','Premio adicional especial de produccion','HR',' ',150,'TRUE', 0,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('1543','Adicional articulo 33','HR',' ',155,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1547','Incentivo a la produccion','HR',' ',160,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1563','Ropa de trabajo','HN',' ',240,'TRUE', 0,0,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('1583','Adicional trabajos especiales','HR',' ',165,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1603','Categoria superior','HR',' ',170,'TRUE', 1,1,1,0,0,1,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('1653','Diferencia remuneracion asegurada','HR',' ',175,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    # PERSONAL TIERRA
    ('1693','Horas normales','HR',' ',10,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('1803','Horas extras 50%','HR',' ',20,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('1813','Horas extras 100%','HR',' ',25,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('1823','Adicional frio','HR',' ',30,'TRUE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1863','A cuenta futuros aumentos (STIA)','HR',' ',35,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1883','Gratificacion trabajos en planta','HR',' ',40,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1893','Horas extras 150%','HR',' ',27,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('1913','Adicional presencia','HR',' ',45,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('1933','Bonificacion unica extraordinaria','HR',' ',250,'FALSE', 1,1,1,0,1,1,1, 0,0,0,0,0,1,0,0,0,0,''),
    ('2003','A cuenta futuros aumentos (175/75)','HR',' ',180,'TRUE', 1,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2013','A cuenta futuros aumentos puerto','HR',' ',182,'TRUE', 1,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2023','A cuenta futuros aumentos dique','HR',' ',184,'TRUE', 1,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2033','A cuenta futuros aumentos pilotaje','HR',' ',186,'TRUE', 1,1,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2043','Gratificacion trabajos en puerto','HR',' ',188,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2053','Gratificacion trabajos en dique','HR',' ',190,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2073','Adicional tareas mecanicas','HR',' ',195,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2083','A cuenta futuros aumentos francos','HR',' ',197,'TRUE', 1,0,0,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2103','Presentismo Comercio','HR',' ',45,'TRUE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,0,1,''),
    ('2303','Porcentaje sobre valor','HR',' ',110,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2313','Diferencia garantizada','HR',' ',115,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2323','Reduccion pactada','HN',' ',245,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2343','Adicional por produccion','HR',' ',120,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('2353','Reduccion pactada 2016','HN',' ',248,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2363','Reduccion pactada 2020','HN',' ',249,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2403','Produccion langostino entero 1','HR',' ',141,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2413','Produccion langostino entero 2','HR',' ',142,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2423','Produccion langostino entero 3','HR',' ',143,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2433','Produccion langostino entero 4','HR',' ',144,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2443','Produccion langostino entero 5','HR',' ',145,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2453','Produccion langostino entero 6','HR',' ',146,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2463','Produccion langostino cola 1','HR',' ',147,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2473','Produccion langostino cola 2','HR',' ',148,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2483','Produccion langostino cola 3','HR',' ',149,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,1,0,''),
    ('2535','Complemento no rem produccion merluza','HN',' ',260,'TRUE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('2553','Retroactivo acuerdo SOMU','HR',' ',270,'FALSE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2573','Ajuste de adicional bodega','HR',' ',275,'FALSE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2603','Ajuste de horas extras 150%','HR',' ',325,'FALSE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('2613','Ajuste de marea','HR',' ',280,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2673','Ajuste','HR',' ',285,'FALSE', 1,1,1,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('2675','Retroactivo incremento paritaria STIA','HR',' ',290,'FALSE', 0,0,0,0,0,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2703','Adicional por zona desfavorable','HR',' ',55,'TRUE', 0,0,0,0,1,1,1, 1,1,1,1,1,1,1,0,0,1,''),
    ('2713','Gratificacion','HR',' ',295,'TRUE', 0,0,0,0,0,0,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('2733','Gratificacion por jubilacion','HR',' ',300,'TRUE', 0,1,1,0,0,0,0, 1,0,0,0,0,0,0,0,0,0,''),
    ('2743','Ajuste de sueldo a ordenes','HR',' ',305,'FALSE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('2763','Ajuste de horas normales','HR',' ',310,'FALSE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('2773','Ajuste de horas extras 50%','HR',' ',315,'FALSE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('2783','Ajuste de horas extras 100%','HR',' ',320,'FALSE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('2803','Indemnizacion francos no gozados','HR','Liquidacion Final',400,'TRUE', 1,1,1,0,0,0,0, 1,0,0,0,0,0,0,0,0,0,''),
    ('2804','SAC s/ francos no gozados','HR','Liquidacion Final',405,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('3473','Gastos de sepelio articulo 84 CCT 372/04','HR',' ',330,'TRUE', 0,0,0,0,0,1,0, 1,0,0,0,0,0,0,0,0,0,''),
    ('3474','Subsidio jubilacion articulo 82 CCT 372/04','HR',' ',335,'TRUE', 0,0,0,0,0,1,0, 1,0,0,0,0,0,0,0,0,0,''),
    ('3553','Vacaciones','HR','Vacaciones',450,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,0,0,0,0,0,''),
    ('3573','Adicional vacaciones articulo 76 CCT 372/04','HR','Vacaciones',455,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,1,0,0,0,0,0,''),
    ('3613','SAC primer semestre','HR','Aguinaldo',460,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,0,0,0,0,0,''),
    ('3623','SAC segundo semestre','HR','Aguinaldo',465,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,0,0,0,0,0,''),
    ('3813','SAC egreso primer semestre','HR','Liquidacion Final',470,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,0,0,0,0,0,''),
    ('3823','SAC egreso segundo semestre','HR','Liquidacion Final',475,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,0,0,0,0,0,''),
    ('3903','Indemnizacion sustitutiva del preaviso','HR','Liquidacion Final',500,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3904','SAC s/ indem. sustitutiva preaviso','HR','Liquidacion Final',505,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('3913','Indemnizacion antiguedad por despido','HR','Liquidacion Final',510,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3923','Indemnizacion antiguedad por fallecimiento','HR','Liquidacion Final',515,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3924','Indemnizacion enfermedad art. 213 LCT','HR','Liquidacion Final',520,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('3933','Indemnizacion incapacidad art. 212 P2 LCT','HR','Liquidacion Final',525,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3935','Indemnizacion incapacidad art. 212 P3 LCT','HR','Liquidacion Final',530,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3937','Indemnizacion incapacidad art. 212 P4 LCT','HR','Liquidacion Final',535,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3943','Integracion mes de despido','HR','Liquidacion Final',540,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3944','SAC s/ integracion mes de despido','HR','Liquidacion Final',545,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('3953','Indemnizacion causa maternidad','HR','Liquidacion Final',550,'TRUE', 0,0,0,0,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3973','Vacaciones no gozadas','HR','Liquidacion Final',555,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3974','SAC s/ vacaciones no gozadas','HR','Liquidacion Final',560,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('3975','Vacaciones no gozadas anios anteriores','HR','Liquidacion Final',565,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3976','SAC s/ vacaciones no gozadas anios ant.','HR','Liquidacion Final',570,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('3983','Indemnizacion fuerza mayor','HR','Liquidacion Final',575,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('3993','Indemnizacion causa embarazo','HR','Liquidacion Final',580,'TRUE', 0,0,0,0,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('4003','Indemnizacion causa matrimonio','HR','Liquidacion Final',585,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('4013','Indemnizacion art. 52 P4 Ley 23551','HR','Liquidacion Final',590,'TRUE', 1,1,1,0,1,1,0, 1,0,0,0,0,0,0,0,0,0,''),
    ('4023','Conciliacion','HR','Liquidacion Final',595,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('4033','Bonificacion fin de relacion laboral','HR','Liquidacion Final',600,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('4060','Licencia gremial','HR',' ',340,'TRUE', 0,0,0,0,1,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4070','Licencia por casamiento','HR',' ',345,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4080','Licencia por nacimiento','HR',' ',350,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4090','Licencia por defuncion','HR',' ',355,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4100','Licencia por examen','HR',' ',360,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4108','Licencia por donacion de sangre','HR',' ',365,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4110','Enfermedad inculpable','HR',' ',370,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4111','Ajuste de enfermedad inculpable','HR',' ',680,'FALSE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4115','Descuento dias de enfermedad','DE',' ',610,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4120','Accidente - Enfermedad profesional','HR',' ',375,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,0,1,0,1,0,0,''),
    ('4121','Descuento dias de accidente','DE',' ',615,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,0,1,0,1,0,0,''),
    ('4123','Horas de accidente - enf. profesional','HR',' ',380,'TRUE', 0,0,0,0,1,0,0, 1,1,1,1,0,1,1,0,0,0,''),
    ('4124','Ajuste prestacion ILT','HR',' ',685,'FALSE', 1,1,1,1,1,1,1, 1,1,1,1,0,1,0,1,0,0,''),
    ('4125','Prestacion ILT','HR',' ',385,'TRUE', 1,1,1,1,1,1,1, 1,1,1,1,0,1,0,1,0,0,''),
    ('4135','Descuento dias de maternidad','DE',' ',620,'TRUE', 0,0,0,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4140','Suspension disciplinaria','DE',' ',625,'TRUE', 1,1,1,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4160','Accidente - Enf. profesional no remunerativo','HN',' ',390,'TRUE', 1,1,1,1,1,1,1, 0,1,1,0,0,1,0,1,0,0,''),
    ('4165','Prestacion ILT no remunerativa','HN',' ',395,'TRUE', 1,1,1,1,1,1,1, 0,1,1,0,0,1,0,1,0,0,''),
    ('4213','Anticipo de vacaciones','HR',' ',630,'TRUE', 0,0,0,0,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('4243','Ajuste de vacaciones','HR',' ',635,'FALSE', 1,1,1,1,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4415','Absorcion articulo 36 CCT SOMU-CAPECA','HR',' ',400,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('4423','Horas extras 100% (729/15)','HR',' ',405,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('4433','Horas extras 50% (729/15)','HR',' ',410,'TRUE', 0,0,1,0,0,0,0, 1,1,1,1,1,1,1,1,0,0,''),
    ('4443','Adicional presentismo STIA','HR',' ',415,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4444','Bonificacion especial remunerativa STIA','HR',' ',420,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4453','Dias feriados','HR',' ',425,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('4473','Ajuste de dias feriados','HR',' ',430,'FALSE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,1,0,''),
    ('4573','Asistencia perfecta','HR',' ',435,'TRUE', 0,0,0,0,1,1,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4680','Dec 194/23 prog incremento exportador','HR',' ',440,'TRUE', 1,1,1,0,0,0,0, 1,1,1,1,1,1,1,0,0,0,''),
    ('4683','Incremento no rem paritaria STIA','HN',' ',445,'TRUE', 0,0,0,0,0,1,0, 0,1,1,0,0,1,0,0,0,0,''),
    ('4684','Retroactivo incremento no rem STIA','HN',' ',450,'FALSE', 0,0,0,0,0,1,0, 0,1,1,0,0,1,0,0,0,0,''),
    ('4703','Deduccion por ausencias','DE',' ',640,'TRUE', 0,0,0,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4723','Deduccion por huelga','DE',' ',645,'TRUE', 0,0,0,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4743','Descuento dias de vacaciones','DE',' ',650,'TRUE', 0,0,0,0,1,1,1, 1,1,1,1,1,1,1,0,0,0,''),
    ('4803','Anticipo de poder','HR',' ',655,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('4805','Adelanto incentivo capacitacion','HR',' ',660,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('4813','Anticipo de llegada','HR',' ',665,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('4823','Anticipo de haberes','HR',' ',670,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('4833','Porcentaje sobre valor (Adelanto haberes)','HR',' ',675,'TRUE', 0,0,0,1,0,0,0, 1,1,1,1,1,0,0,0,0,0,''),
    ('4853','Asignacion extraord. no rem Comercio','HN',' ',455,'TRUE', 0,0,0,0,1,0,0, 0,1,1,0,0,1,0,0,0,0,''),
    ('4856','Incremento no rem acuerdo Comercio','HN',' ',458,'TRUE', 0,0,0,0,1,0,0, 1,1,1,0,0,1,0,0,0,0,''),
    ('4873','Bonificacion extraordinaria SICONARA','HR',' ',460,'TRUE', 1,0,0,0,0,0,0, 0,1,1,0,0,1,0,0,0,0,''),
    ('4893','Incremento no rem acuerdo SOMU','HN',' ',462,'TRUE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('4894','Retroactivo incremento no rem SOMU','HN',' ',464,'FALSE', 0,0,1,0,0,0,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('4895','Incremento no rem paritaria STIA (2)','HN',' ',466,'TRUE', 0,0,0,0,0,1,0, 1,1,1,1,0,1,0,0,0,0,''),
    ('4896','Retroactivo incremento no rem STIA (2)','HN',' ',468,'FALSE', 0,0,0,0,0,1,0, 0,1,1,1,0,1,0,0,0,0,''),
    ('5498','Redondeo','HR',' ',999,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    # RETENCIONES
    ('5010','Impuesto a las ganancias','RE',' ',850,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"TRAMO('TAB_IMP_4CAT',MAX(BASE_IG4*12-MNI_ANUAL,0))/12"),
    ('5310','Impuesto a las ganancias anio anterior','RE',' ',855,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8522','Cuota sindical','DE',' ',700,'TRUE', 1,1,1,0,1,1,0, 0,0,0,0,0,0,0,0,0,0,'BASE_SINDICAL * PCT_CUOTA_SIND'),
    ('8523','Cuota sindical SAC','DE','Aguinaldo',705,'TRUE', 1,1,1,0,1,1,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8525','Cuota sindical vacaciones','DE','Vacaciones',710,'TRUE', 1,1,1,0,1,1,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8542','Caja compensadora','DE',' ',715,'TRUE', 0,1,0,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8543','Fondo de desempleo','DE',' ',720,'TRUE', 1,1,0,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8553','Centro de capacitacion','DE',' ',725,'TRUE', 0,1,0,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8563','Contribucion solidaria SOMU','DE',' ',730,'TRUE', 0,0,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8573','Aporte solidario STIA','DE',' ',735,'TRUE', 0,0,0,0,0,1,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8593','Cuota extraordinaria accion social CAP','DE',' ',740,'TRUE', 0,1,0,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8603','Contribucion solidaria AACPyPP','DE',' ',745,'TRUE', 0,1,0,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8623','Gastos de hotel','DE',' ',750,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8633','Gastos de farmacia','DE',' ',755,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8643','Gastos de pasajes','DE',' ',760,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8653','Gastos medicos','DE',' ',765,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8683','Gastos de cigarrillos','DE',' ',770,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8693','Descuento de comunicaciones','DE',' ',775,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8703','Gastos varios','DE',' ',780,'FALSE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8810','Descuento anticipo vacaciones','DE',' ',785,'TRUE', 0,0,0,0,1,1,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8830','Descuento anticipo poder','DE',' ',790,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8835','Descuento adelanto incentivo capacitacion','DE',' ',795,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8840','Descuento anticipo llegada','DE',' ',800,'TRUE', 1,1,1,0,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8850','Descuento anticipo haberes','DE',' ',805,'TRUE', 1,1,1,1,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8860','Porcentaje sobre valor (Descuento adelanto)','DE',' ',810,'TRUE', 0,0,0,1,0,0,0, 0,0,0,0,0,0,0,0,0,0,''),
    ('8870','Descuento anticipo haberes ADM','DE',' ',815,'TRUE', 0,0,0,0,1,0,1, 0,0,0,0,0,0,0,0,0,0,''),
    ('8883','Embargo judicial','RE',' ',820,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    ('8893','Cuota alimentaria','RE',' ',825,'TRUE', 1,1,1,1,1,1,1, 1,0,0,0,0,0,0,0,0,0,''),
    # SEGURIDAD SOCIAL
    ('6000','Jubilacion','SS',' ',900,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SS,TOPE_SIPA)*0,11"),
    ('6002','Jubilacion SAC','SS','Aguinaldo',902,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA)*0,11"),
    ('6003','Jubilacion vacaciones','SS','Vacaciones',904,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA)*0,11"),
    ('6010','Ley 19032','SS',' ',906,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SS,TOPE_SIPA)*0,03"),
    ('6012','Ley 19032 SAC','SS','Aguinaldo',908,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA)*0,03"),
    ('6013','Ley 19032 vacaciones','SS','Vacaciones',910,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA)*0,03"),
    ('6030','Obra social','SS',' ',912,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_OS,TOPE_SIPA_OS)*0,03"),
    ('6032','Adicional obra social','SS',' ',914,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_OS,TOPE_SIPA_OS)*0,015"),
    ('6034','Obra social SAC','SS','Aguinaldo',916,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA_OS)*0,03"),
    ('6035','Obra social vacaciones','SS','Vacaciones',918,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA_OS)*0,03"),
    ('6036','Adicional obra social SAC','SS','Aguinaldo',920,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA_OS)*0,015"),
    ('6037','Adicional obra social vacaciones','SS','Vacaciones',922,'TRUE', 1,1,1,1,1,1,1, 0,0,0,0,0,0,0,0,0,0,"MIN(BASE_SAC,TOPE_SIPA_OS)*0,015"),
]

def hdr(cell, bg='1F497D'):
    cell.font = Font(bold=True, color='FFFFFF', size=9)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

def sheet(wb, name, headers, rows, bg='1F497D'):
    ws = wb.create_sheet(name[:31])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        hdr(cell, bg)
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(h))+2)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    ws.freeze_panes = 'A2'
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── 1. Instrucciones ─────────────────────────────────────────────────
ws = wb.create_sheet('INSTRUCCIONES')
ws.column_dimensions['A'].width = 80
instrucciones = [
    ['ConfigPackage_Payroll.xlsx — Paquete de configuracion para extension Payroll BC'],
    [''],
    ['COMO IMPORTAR EN BC:'],
    ['1. Abrir Paquetes de configuracion (Config. Packages)'],
    ['2. Crear nuevo paquete con un codigo (ej. PAYROLL-25)'],
    ['3. Importar este Excel via accion "Importar desde Excel"'],
    ['4. Validar cada tabla y aplicar'],
    [''],
    ['ORDEN DE APLICACION RECOMENDADO:'],
    ['  1. Convenio Colectivo'],
    ['  2. Parametro + Parametro Vigente'],
    ['  3. Variable Sistema Liq.'],
    ['  4. Concepto Liquidacion (incluye acumuladores)'],
    ['  5. Concepto CCT Vigente'],
    ['  6. Fraccion Acumulador'],
    [''],
    ['NOTAS:'],
    ['- Completar Parametro Vigente con valores reales antes de calcular'],
    ['- Completar Categoria CCT con escalas de cada convenio'],
    ['- Los conceptos con Activo=FALSE son cerrados o no implementados aun'],
    ['- Formula es orientativa; revisar y completar segun CCT'],
    ['- Fraccion Acumulador: el % debe sumar 100 por concepto+vigencia'],
    ['- Concepto CCT Vigente: si no hay filas para un concepto, aplica a TODOS los CCT'],
]
for r, row in enumerate(instrucciones, 1):
    ws.cell(r, 1, row[0] if row else '')
ws.cell(1,1).font = Font(bold=True, size=12)

# ── 2. Convenio Colectivo ────────────────────────────────────────────
ccts_data = [
    ('175/75','CCT 175/75 Personal Embarcado Oficiales CAPECA','','',''),
    ('768/19','CCT 768/19 Personal Embarcado Oficiales AACPyPP','','',''),
    ('729/15','CCT 729/15 Personal Embarcado Marineria SOMU-CAPECA','','',''),
    ('ESP','Sin CCT - Personal Embarcado Espania','','',''),
    ('130/75','CCT 130/1975 Comercio','','',''),
    ('372/04','CCT 372/2004 STIA','','',''),
    ('ADM','Sin CCT - Personal Administrativo','','',''),
]
sheet(wb,'Convenio Colectivo',
    ['Codigo','Descripcion','No. CCT','Sindicato','Observaciones'],
    ccts_data, '2E75B6')

# ── 3. Parametro ─────────────────────────────────────────────────────
params_data = [
    ('BASICO','Basico del convenio/categoria','Clave efectiva: BASICO_[CCT]_[CAT]','BASICO','TRUE'),
    ('SMVM','Salario Minimo Vital y Movil','','SMVM','FALSE'),
    ('TC_COMPRADOR','Tipo de cambio comprador','','TC_COMPRADOR','FALSE'),
    ('TOPE_SIPA','Tope imponible SIPA (remunerativo)','','TOPE_SIPA','FALSE'),
    ('TOPE_SIPA_OS','Tope imponible SIPA (obra social)','','TOPE_SIPA_OS','FALSE'),
    ('MNI_ANUAL','Minimo no imponible anual Ganancias 4ta','','MNI_ANUAL','FALSE'),
    ('PCT_ANTIG','Porcentaje de antiguedad por anio','Ej: 0,01 = 1% por anio','PCT_ANTIG','FALSE'),
    ('PCT_CUOTA_SIND','Tasa cuota sindical','Varia por CCT','PCT_CUOTA_SIND','FALSE'),
]
sheet(wb,'Parametro',
    ['Codigo','Descripcion','Notas','Nombre Variable','Sufijo CCT'],
    params_data, '2E75B6')

# ── 4. Parametro Vigente ─────────────────────────────────────────────
pv_data = [
    ('SMVM',         VIG, 'Vigente '+VIG, 0, '', 'FALSE', ''),
    ('TC_COMPRADOR', VIG, 'Vigente '+VIG, 0, 'USD', 'FALSE', ''),
    ('TOPE_SIPA',    VIG, 'Vigente '+VIG, 0, '', 'FALSE', ''),
    ('TOPE_SIPA_OS', VIG, 'Vigente '+VIG, 0, '', 'FALSE', ''),
    ('MNI_ANUAL',    VIG, 'Vigente '+VIG, 0, '', 'FALSE', ''),
    ('PCT_ANTIG',    VIG, 'Vigente '+VIG, 0.01, '', 'FALSE', ''),
    ('PCT_CUOTA_SIND',VIG,'Vigente '+VIG, 0, '', 'FALSE', 'Completar por CCT'),
]
sheet(wb,'Parametro Vigente',
    ['Cod. Parametro','Vigencia Desde','Descripcion Version','Valor','Moneda','En Uso','Notas'],
    pv_data, '2E75B6')

# ── 5. Variable Sistema Liq. ─────────────────────────────────────────
vs_data = [
    ('ANIOS_ANTIGUEDAD','ANIOS_ANTIGUEDAD','Anios de antiguedad del empleado','TRUE'),
    ('DIAS_HAB','DIAS_HAB','Dias habiles (Lun-Vie) del periodo','TRUE'),
    ('PCT_ESCALA','PCT_ESCALA','Escala % de la categoria CCT / 100','TRUE'),
    ('DIAS_PROYECTO','DIAS_MAR','Dias del proyecto (Fecha Fin - Fecha Inicio)','TRUE'),
]
sheet(wb,'Variable Sistema Liq.',
    ['Cod. Calculo','Nombre Variable','Descripcion','Activo'],
    vs_data, '2E75B6')

# ── 6. Concepto Liquidacion (acumuladores + conceptos) ───────────────
conc_headers = [
    'Codigo','Vigencia Desde','Descripcion','Nombre Impresion',
    'Tipo Concepto','Formula','Condicion','Orden Calculo',
    'Aplica A','Activo','Es Acumulador','Aplica Tipo Liq.'
]
conc_rows = []
# Acumuladores primero
for cod, desc, tipo in ACUMULADORES:
    conc_rows.append((cod, VIG, desc, '', tipo, '', '', 0, 'Todos', 'TRUE', 'TRUE', ' '))

# Conceptos normales
for c in CONCEPTOS:
    cod,nombre,tipo_k,aplica_tipo,ord_c,activo = c[0],c[1],c[2],c[3],c[4],c[5]
    formula = c[21] if len(c)>21 else ''
    tipo_desc = TIPO_MAP[tipo_k]
    conc_rows.append((
        cod, VIG, nombre, nombre[:50],
        tipo_desc, formula, '', ord_c,
        'Todos', activo, 'FALSE', aplica_tipo
    ))

ws_conc = sheet(wb,'Concepto Liquidacion', conc_headers, conc_rows, '375623')
# Color rows by type
tipo_colors = {
    'Haber Remunerativo': 'E2EFDA',
    'Haber No Remunerativo': 'EBF3E8',
    'Descuento Empleado': 'FCE4D6',
    'Retencion': 'FCE4D6',
    'Seguridad Social': 'FFF2CC',
    'Contribucion Patronal': 'DAEEF3',
}
for r in range(2, len(conc_rows)+2):
    tipo_val = ws_conc.cell(r,5).value or ''
    color = tipo_colors.get(tipo_val, 'FFFFFF')
    for c in range(1, len(conc_headers)+1):
        ws_conc.cell(r,c).fill = PatternFill('solid', fgColor=color)

# ── 7. Concepto CCT Vigente ───────────────────────────────────────────
cct_rows = []
for c in CONCEPTOS:
    cod = c[0]
    cct_flags = list(c[6:13])   # 7 CCT flags
    # Only create rows if NOT all CCTs apply (if all, leave empty = applies to all)
    applies_to = [CCT_CODES[i] for i,f in enumerate(cct_flags) if f==1]
    if len(applies_to) < len(CCT_CODES):  # not universal
        for cct in applies_to:
            cct_rows.append((cod, VIG, cct))

sheet(wb,'Concepto CCT Vigente',
    ['Cod. Concepto','Vigencia Desde','Cod. Convenio'],
    cct_rows, '7030A0')

# ── 8. Fraccion Acumulador ────────────────────────────────────────────
frac_rows = []
for c in CONCEPTOS:
    cod, tipo_k = c[0], c[2]
    acc_flags = list(c[13:23])  # 10 accumulator flags

    # Primary accumulator (100%)
    prim = PRIMARY_ACC.get(tipo_k)
    if prim:
        frac_rows.append((cod, VIG, prim, 100, 'Acumulador principal'))

    # Base accumulators (100% each, only for earnings)
    if tipo_k in ('HR', 'HN'):
        for i, flag in enumerate(acc_flags):
            if flag:
                frac_rows.append((cod, VIG, ACC_NAMES[i], 100, f'Base {ACC_NAMES[i]}'))

sheet(wb,'Fraccion Acumulador',
    ['Cod. Concepto','Vigencia Desde','Cod. Acumulador','Porcentaje','Descripcion'],
    frac_rows, '7030A0')

# ── 9. Tabla Escalonada (placeholder) ────────────────────────────────
sheet(wb,'Tabla Escalonada',
    ['Codigo','Vigencia Desde','Descripcion'],
    [('TAB_IMP_4CAT', VIG, 'Impuesto Ganancias 4ta Categoria')], 'C55A11')

sheet(wb,'Tabla Escalonada Det.',
    ['Codigo','Vigencia Desde','No. Tramo','Limite Inferior','Limite Superior','Monto Fijo','Porcentaje','Descripcion'],
    [('TAB_IMP_4CAT', VIG, 1, 0, 0, 0, 0, 'Completar con tabla AFIP vigente')], 'C55A11')

wb.save(OUTPUT)
print(f'Generado: {OUTPUT}')
print(f'  Convenios CCT:      {len(ccts_data)} filas')
print(f'  Parametros:         {len(params_data)} filas')
print(f'  Variables Sistema:  {len(vs_data)} filas')
print(f'  Conceptos:          {len(conc_rows)} filas ({len(ACUMULADORES)} acum + {len(CONCEPTOS)} conceptos)')
print(f'  Concepto CCT:       {len(cct_rows)} filas')
print(f'  Fraccion Acumulador:{len(frac_rows)} filas')
